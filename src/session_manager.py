# src/session_manager.py
"""
Per-session bookkeeping and output writing.

All files for one session live inside ONE folder created by the launcher
and passed via the SESSION_FOLDER environment variable:

    data/session_<YYYYMMDD>_<HHMMSS>_<patient_id>_s<n>/
        ├── metadata.json    (intake + frozen baseline merged)
        ├── samples.csv      (50 Hz clinical record — main artifact)
        ├── unity_udp.csv    (UDP audit log)
        └── diagnostic.csv   (raw + smoothed signal trace, per tick)

Responsibilities of SessionManager:

  * Read the launcher-written metadata.json (intake portion) so every CSV
    row carries the full patient demographics.
  * Maintain samples.csv (the canonical 50 Hz record). Rotated to
    samples_002.csv / samples_003.csv on each LIVE_RESTART so multiple live
    runs against the same baseline land in distinct files inside the same
    session folder.
  * At baseline lock, APPEND the frozen baseline to metadata.json
    (replacing the placeholder `"baseline": null` written by the launcher).
  * Maintain diagnostic.csv: one row per tick, written from main.py via
    log_diagnostic_row(). Acquisition + processing no longer write their
    own diagnostic CSVs — they hand the values to main, main passes them
    here. Single source of truth for per-tick diagnostic data.
  * Track in-memory per-session history (signal_history, stress_history)
    for the print_session_summary at LIVE → STOPPED.
  * Provide cleanup helpers for the shutdown-discard paths.
"""

import json
import os
import time
from datetime import datetime

from config import Config


# ---------- module-level helpers --------------------------------------------

def _session_folder_from_env() -> str:
    """
    Return the per-session folder path the launcher created. Falls back to
    a fresh folder under data/ when SESSION_FOLDER isn't set (e.g. tests
    that import this module without running through the launcher).
    """
    sf = os.environ.get('SESSION_FOLDER')
    if sf and os.path.isdir(sf):
        return sf
    # Fallback for standalone testing.
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(project_root, 'data')
    os.makedirs(data_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fallback = os.path.join(data_dir, f"session_{ts}_TEST_s1")
    os.makedirs(fallback, exist_ok=True)
    return fallback


def _load_metadata_from_env() -> dict:
    """
    Read the launcher-written metadata.json from the session folder.
    Returns the full dict; the 'patient' subkey carries demographics and
    'baseline' is None until write_baseline_capture() fills it in.
    """
    sf = _session_folder_from_env()
    path = os.path.join(sf, 'metadata.json')
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    # Minimal fallback so a stray test import doesn't crash.
    return {
        'patient': {
            'first_name':     os.environ.get('PATIENT_NAME', 'PATIENT'),
            'last_name':      '',
            'patient_id':     os.environ.get('PATIENT_ID', '000'),
            'gender':         '',
            'session_date':   datetime.now().strftime('%Y-%m-%d'),
            'session_number': 1,
        },
        'session_folder': os.path.basename(sf),
        'session_timestamp': datetime.now().strftime("%Y%m%d_%H%M%S"),
        'baseline': None,
    }


# ---------- main class -------------------------------------------------------

def _coerce_xlsx_cell(s: str):
    """Best-effort string -> int/float for write_samples_xlsx, so Excel
    treats numeric columns as numbers (sortable, plottable). Empty cells
    stay empty so they map to Excel's blank semantics."""
    if s == "" or s is None:
        return None
    try:
        if '.' in s or 'e' in s or 'E' in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return s


class SessionManager:
    """Centralised session state + per-tick file writing."""

    # samples.csv header — first column is `sample_n` (a 1-indexed row
    # counter), not a wall-clock timestamp. Each call to log_sample()
    # increments an internal call counter; rows are written every
    # `_samples_write_every_n` calls (50 calls at 50 Hz pipeline rate
    # = one written row per second). The sample_n column gives a clean
    # monotonic index for plotting / aligning across files without the
    # visual clutter of timestamps.
    # Column groups, in the order they appear in the file:
    #   1. row index               sample_n
    #   2. person info             patient_first_name ... session_number
    #   3. status                  phase, state, dashboard_score
    #   4. signals                 eda, hr, hrv, delta_eda, delta_hr, delta_hrv,
    #                              s_instant, s_t
    #   5. scenario telemetry      <dynamic columns learned from the first
    #                              non-empty telemetry packet Unity sends>
    #   6. diagnostic counters     artifacts_eda, artifacts_hr, artifacts_hrv
    #
    # Group 5 is DYNAMIC. Biofeedback is passive: whatever `data` fields
    # the connected Unity scene emits become the CSV columns. Nothing is
    # pre-declared or configured — the scene ships one JSON packet per
    # tick with an arbitrary `data` object, and its keys turn into
    # samples.csv columns in the order Unity sends them.
    #
    # Because Baseline rows are written BEFORE Unity Play (no telemetry
    # yet), those rows are buffered in memory. As soon as the first
    # non-empty telemetry packet arrives, its keys become the column
    # list, the header is written, and the buffered rows are flushed
    # with empty telemetry cells. If a session ends without any
    # telemetry (e.g. physiology-only test), the header is committed
    # with no dynamic columns and the buffer is flushed anyway.
    #
    # Once committed, the column set stays fixed for the whole session.
    # Fields Unity subsequently adds are dropped (with a one-time
    # console warning); missing fields land as empty cells.
    _SAMPLES_HEADER_PREFIX = ("sample_n,"
                              "patient_first_name,patient_last_name,patient_id,"
                              "gender,session_date,session_number,"
                              "phase,state,dashboard_score,"
                              "eda,hr,hrv,"
                              "delta_eda,delta_hr,delta_hrv,s_instant,s_t")
    _SAMPLES_HEADER_SUFFIX = "artifacts_eda,artifacts_hr,artifacts_hrv"

    # diagnostic.csv header — `tick_n` is a 1-indexed counter of rows
    # actually written (not pipeline ticks; rows are decimated per
    # Config.DIAGNOSTIC_CSV_RATE_HZ). Pipeline frequency itself is
    # unchanged (50 Hz); only the disk-write rate is throttled.
    # The old smooth_eda/hr/hrv columns were dropped — with EMA removed
    # per PDF, "smoothed" was identical to "raw" so they were just clutter.
    _DIAGNOSTIC_HEADER = ("tick_n,phase,status,"
                          "raw_eda,raw_hr,raw_hrv\n")

    def __init__(self, patient_name: str = "PATIENT", patient_id: str = "000"):
        # Load metadata from the session folder (launcher wrote it).
        self.metadata = _load_metadata_from_env()
        self.patient = self.metadata['patient']

        # Allow positional args to override (compat with main.py's old calls).
        if patient_name != "PATIENT":
            self.patient['first_name'] = patient_name
        if patient_id != "000":
            self.patient['patient_id'] = patient_id

        self.patient_name = self.patient['first_name']
        self.patient_id = self.patient['patient_id']
        self.patient_info = f"{self.patient_name}_{self.patient_id}"

        # Dynamic-column bookkeeping. Determined at runtime from the
        # first non-empty telemetry packet Unity sends; None until then.
        self._dynamic_fields = None          # ordered list of field names
        self._header_committed = False       # True once samples.csv header is written
        self._observed_scenario = ""         # remembered for metadata.json
        # Rows that arrived before the first telemetry packet — held here
        # until the header can be written, then flushed with empty cells
        # in the dynamic columns.
        self._pending_rows = []
        # Track unknown telemetry fields we've already warned about, so
        # runaway Unity streams don't spam the console.
        self._unknown_field_warnings = set()

        # Session folder + per-file paths inside it.
        self.session_folder = _session_folder_from_env()
        self.metadata_path = os.path.join(self.session_folder, 'metadata.json')
        self.samples_path = os.path.join(self.session_folder, 'samples.csv')
        self.diagnostic_path = os.path.join(self.session_folder, 'diagnostic.csv')
        # unity_udp.csv is written by UnityUDPBridge — main.py passes the
        # path; we just expose where it lives for the summary print.
        self.unity_udp_path = os.path.join(self.session_folder, 'unity_udp.csv')

        self.session_start_time = time.time()
        self.session_start_datetime = datetime.now()
        self.session_timestamp = self.metadata.get(
            'session_timestamp',
            self.session_start_datetime.strftime("%Y%m%d_%H%M%S"),
        )

        # Counter-based decimation for log_sample() and log_diagnostic_row().
        # We write one row every Nth call where
        #     N = PIPELINE_RATE / target_rate.
        # At 50 Hz pipeline / 1 Hz target → write every 50 calls → ~1 row/sec.
        # Pipeline frequency is unchanged; only the disk write is throttled.
        #
        # Two counters per file:
        #   _<file>_call_count : every call to log_*() increments this.
        #     Used for the modulo gate (skip if not at the threshold).
        #   _<file>_row_n      : increments only when a row is actually
        #     written. Goes into the sample_n / tick_n column for a clean
        #     monotonic index in the file.
        self._samples_write_every_n = max(
            1, round(Config.PIPELINE_RATE / max(Config.SAMPLES_CSV_RATE_HZ, 1e-6)))
        self._diagnostic_write_every_n = max(
            1, round(Config.PIPELINE_RATE / max(Config.DIAGNOSTIC_CSV_RATE_HZ, 1e-6)))
        self._samples_call_count = 0
        self._diagnostic_call_count = 0
        self._samples_row_n = 0
        self._diagnostic_row_n = 0

        # Phase tracking, signal history, stress history — unchanged shape so
        # everything that reads these (dashboard summary, session_review)
        # keeps working.
        self.phase = "BASELINE"
        self.baseline_end_time = None
        self.baseline_ticks_remaining = int(Config.BASELINE_SEC * Config.PIPELINE_RATE)
        self.signal_history = {'eda': [], 'hr': [], 'hrv': [], 'timestamps': []}
        self.max_history_length = int(Config.PIPELINE_RATE * 60)
        self.personal_baselines = None
        self.baseline_buffers = None
        self.artifacts_removed = {'eda': 0, 'hr': 0, 'hrv': 0}
        self.stress_history = {
            's_instant': [], 's_t': [], 'state': [], 'dashboard_score': [],
        }
        self.thresh_mild = None
        self.thresh_high = None

        # Open file handles for the lifetime of the session.
        # samples.csv opens WITHOUT the header — the header is deferred
        # until we have seen the first telemetry packet from Unity and
        # can list its `data` fields as dynamic columns. All log_sample
        # calls before that first packet buffer their rows in memory.
        self._samples_handle = None
        self._diagnostic_handle = None
        self._open_samples(write_header=False)
        self._open_diagnostic(write_header=True)

        # Expose canonical paths for callers (main.py uses this for the
        # unity_udp.csv path, summary printing, etc.).
        self.output_file_path = self.samples_path  # back-compat alias

        bar = '=' * 58
        print(f"\n[SESSION] {bar}")
        print(f"[SESSION] New session started")
        print(f"[SESSION]   Patient            : {self.patient_info}")
        print(f"[SESSION]   Start time         : "
              f"{self.session_start_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"[SESSION]   Session folder     : {os.path.basename(self.session_folder)}")
        print(f"[SESSION]   Phase              : {self.phase}")
        print(f"[SESSION]   Baseline duration  : "
              f"{Config.BASELINE_SEC}s ({self.baseline_ticks_remaining} ticks)")
        print(f"[SESSION] {bar}\n")

    # ---------- file-handle helpers -----------------------------------------

    def _build_samples_header(self) -> str:
        """
        Assemble the samples.csv header string:
            <fixed prefix>,<dynamic field 1>,...,<dynamic field N>,<fixed suffix>\n

        Uses whatever `_dynamic_fields` we learned from the first Unity
        packet. If the session never saw a packet, the dynamic slot is
        empty and the header collapses to prefix + suffix.
        """
        parts = [self._SAMPLES_HEADER_PREFIX]
        if self._dynamic_fields:
            parts.extend(self._dynamic_fields)
        parts.append(self._SAMPLES_HEADER_SUFFIX)
        return ",".join(parts) + "\n"

    def _open_samples(self, write_header: bool):
        self._close_samples()
        self._samples_handle = open(self.samples_path, 'w',
                                    newline='', buffering=1)
        if write_header:
            self._samples_handle.write(self._build_samples_header())
            self._samples_handle.flush()

    def _close_samples(self):
        try:
            if self._samples_handle is not None:
                self._samples_handle.flush()
                self._samples_handle.close()
        except Exception:
            pass
        self._samples_handle = None

    def _open_diagnostic(self, write_header: bool):
        self._close_diagnostic()
        self._diagnostic_handle = open(self.diagnostic_path, 'w',
                                       newline='', buffering=1)
        if write_header:
            self._diagnostic_handle.write(self._DIAGNOSTIC_HEADER)
            self._diagnostic_handle.flush()

    def _close_diagnostic(self):
        try:
            if self._diagnostic_handle is not None:
                self._diagnostic_handle.flush()
                self._diagnostic_handle.close()
        except Exception:
            pass
        self._diagnostic_handle = None

    def close_all(self):
        """Close both file handles. Safe to call repeatedly.

        If the session ended without a telemetry packet ever arriving
        (no Unity, or physiology-only test), commit the header with no
        dynamic columns first so the pending baseline rows still land
        on disk instead of being silently dropped.
        """
        if not self._header_committed and self._samples_handle is not None:
            self._commit_header_and_flush()
        self._close_samples()
        self._close_diagnostic()

    # ---------- LIVE_RESTART rotation ---------------------------------------

    def rotate_session_csv(self):
        """
        Called when the operator clicks Start Live a second time after a
        Stop. Per user policy: DELETE the previous live run's samples
        and start fresh. We truncate samples.csv (rewrite the header),
        clear in-memory stress history, and reset the decimation counter
        + sample_n counter so the new run starts at sample_n=1.

        Diagnostic.csv (the forensic raw-signal trace) is NOT truncated
        — it accumulates across the whole launch for post-hoc debugging.
        The baseline (in metadata.json) is untouched: it's still valid.
        """
        self.stress_history = {
            's_instant': [], 's_t': [], 'state': [], 'dashboard_score': [],
        }
        # Fresh file, defer header until the first LIVE-restart telemetry
        # packet arrives (matches the session-start behaviour).
        self._dynamic_fields = None
        self._header_committed = False
        self._pending_rows = []
        self._unknown_field_warnings.clear()
        self._open_samples(write_header=False)
        # Reset both counters so the new run starts at sample_n=1 and the
        # decimation gate fires on its first call.
        self._samples_call_count = 0
        self._samples_row_n = 0
        print("[SESSION] Live restart: previous samples.csv discarded, "
              "fresh recording starting.")

    # Alias retained for any caller that still uses the old name.
    flush_live_data = rotate_session_csv

    # ---------- per-tick writes ---------------------------------------------

    def log_sample(self, eda: float, hr: float, hrv: float,
                   delta_eda: float = 0.0, delta_hr: float = 0.0, delta_hrv: float = 0.0,
                   s_instant: float = None, s_t: float = None,
                   state: str = None, dashboard_score: float = None,
                   height_m=None,        # accepted + ignored (back-compat)
                   scenario: str = "", telemetry_data: dict = None):
        """
        Append one row to samples.csv (held handle, single fwrite).

        Counter-based decimation: called every pipeline tick, but only
        every `_samples_write_every_n`-th call actually writes a row.
        At 50 Hz pipeline / 1 Hz target → one row per 50 calls →
        ~1 row/sec on disk. The pipeline math, the LSL stream, and the
        dashboard are unaffected — they all keep running at 50 Hz.

        First column is `sample_n` (1-indexed row counter), not a
        timestamp. Per user request: the CSV should be readable without
        a noisy time column, and a counter is enough to align rows
        across files and reason about elapsed time (sample_n / rate).
        """
        if self._samples_handle is None:
            return
        self._samples_call_count += 1
        if self._samples_call_count % self._samples_write_every_n != 0:
            return  # decimated — skip this tick
        self._samples_row_n += 1

        # ---- First non-empty telemetry: commit the header now. ----
        # The dynamic columns are the keys of Unity's `data` object, in
        # the order Unity sends them. `scenario` is remembered for
        # metadata.json but does NOT become a CSV column.
        if telemetry_data and not self._header_committed:
            self._dynamic_fields = list(telemetry_data.keys())
            if scenario:
                self._observed_scenario = scenario.strip().lower()
            self._commit_header_and_flush()

        # ---- Warn once per unknown late-arriving field ----
        # If Unity starts adding new fields after the header has been
        # committed, they get dropped. Log the first offender per field
        # so drift between scenes shows up in the console instead of
        # silently disappearing from the CSV.
        if telemetry_data and self._header_committed:
            for fname in telemetry_data:
                if (fname not in self._dynamic_fields
                        and fname not in self._unknown_field_warnings):
                    self._unknown_field_warnings.add(fname)
                    print(f"[SESSION] WARN: telemetry field {fname!r} "
                          f"appeared AFTER the header was committed. It "
                          f"will not be written to samples.csv this "
                          f"session — restart the session to include it.")

        # ---- Build the row (independent of whether we write or buffer) ----
        row = self._format_row(
            eda=eda, hr=hr, hrv=hrv,
            delta_eda=delta_eda, delta_hr=delta_hr, delta_hrv=delta_hrv,
            s_instant=s_instant, s_t=s_t,
            state=state, dashboard_score=dashboard_score,
            telemetry_data=telemetry_data,
        )

        if self._header_committed:
            self._samples_handle.write(row)
        else:
            # Buffer until the first telemetry packet — or until close_all
            # flushes with no dynamic columns if none ever arrives.
            self._pending_rows.append(row)

    def _format_row(self, *, eda, hr, hrv, delta_eda, delta_hr, delta_hrv,
                     s_instant, s_t, state, dashboard_score,
                     telemetry_data):
        """
        Build one samples.csv row string. Dynamic-column cells use the
        current `_dynamic_fields`, formatting each value out of
        `telemetry_data`. Missing / non-registered fields become empty
        cells; unknown late-arriving fields are ignored.
        """
        p = self.patient
        cell_values = []
        for fname in (self._dynamic_fields or []):
            if not telemetry_data or fname not in telemetry_data:
                cell_values.append("")
                continue
            v = telemetry_data[fname]
            try:
                fv = float(v)
                # bool first because bool is a subtype of int in Python.
                if isinstance(v, bool):
                    cell_values.append("1" if v else "0")
                elif isinstance(v, int) or fv.is_integer():
                    cell_values.append(f"{int(fv)}")
                else:
                    cell_values.append(f"{fv:.3f}")
            except (TypeError, ValueError):
                # Non-numeric value — keep it as-is, CSV-quoted.
                s = str(v).replace('"', '""')
                cell_values.append(f'"{s}"')

        dyn = ",".join(cell_values)
        # Comma-join yields "" when there are zero fields; guard against
        # a collapsed double-comma in that case.
        if dyn:
            dyn += ","

        return (
            f"{self._samples_row_n},"
            f"{p['first_name']},{p['last_name']},{p['patient_id']},"
            f"{p['gender']},{p['session_date']},{p['session_number']},"
            f"{self.phase},"
            f"{state if state else 'unknown'},"
            f"{dashboard_score if dashboard_score is not None else 0.0:.4f},"
            f"{eda:.6f},{hr:.6f},{hrv:.6f},"
            f"{delta_eda:.6f},{delta_hr:.6f},{delta_hrv:.6f},"
            f"{s_instant if s_instant is not None else 0.0:.6f},"
            f"{s_t if s_t is not None else 0.0:.6f},"
            f"{dyn}"
            f"{self.artifacts_removed.get('eda', 0)},"
            f"{self.artifacts_removed.get('hr', 0)},"
            f"{self.artifacts_removed.get('hrv', 0)}\n"
        )

    def _commit_header_and_flush(self):
        """
        Write the samples.csv header using the (now-known) dynamic-field
        list and flush every buffered row to disk. Idempotent — safe to
        call more than once; second call is a no-op.
        """
        if self._header_committed:
            return
        if self._samples_handle is None:
            return
        self._samples_handle.write(self._build_samples_header())
        for row in self._pending_rows:
            self._samples_handle.write(row)
        self._samples_handle.flush()
        n = len(self._pending_rows)
        self._pending_rows = []
        self._header_committed = True
        cols = self._dynamic_fields or []
        print(f"[SESSION] samples.csv header committed with "
              f"{len(cols)} dynamic column(s): {cols}. "
              f"Flushed {n} buffered row(s).")

    def log_diagnostic_row(self, status: str,
                            raw_eda: float, raw_hr: float, raw_hrv: float):
        """
        Append one row to diagnostic.csv. Called by main.py once per tick
        with the raw (from acquisition) and smoothed (from processing)
        values plus the acquisition status code (NEW_DATA / HOLD_LAST /
        etc.).

        Counter-based decimation, identical scheme to log_sample. Pipeline
        runs at 50 Hz; this file writes ~1 row/sec on disk (configurable
        via Config.DIAGNOSTIC_CSV_RATE_HZ). First column is `tick_n`,
        a 1-indexed counter of rows written.

        NaN-friendly: any of the float args can be NaN (HR/HRV warm-up
        sentinels per PDF §3). We write "nan" — pandas / Excel both
        understand it on read.
        """
        if self._diagnostic_handle is None:
            return
        self._diagnostic_call_count += 1
        if self._diagnostic_call_count % self._diagnostic_write_every_n != 0:
            return  # decimated — skip this tick
        self._diagnostic_row_n += 1

        def _f(x):
            # 6 decimal places matches log_sample. Preserves small EDA
            # values exactly so the diagnostic and clinical CSVs agree
            # cell-by-cell to the same precision.
            try:
                return f"{x:.6f}"
            except Exception:
                return "nan"

        self._diagnostic_handle.write(
            f"{self._diagnostic_row_n},{self.phase},{status},"
            f"{_f(raw_eda)},{_f(raw_hr)},{_f(raw_hrv)}\n"
        )

    # ---------- in-memory history (for dashboard / summary) -----------------

    def record_raw_sample(self, eda: float, hr: float, hrv: float):
        """Append to the in-memory rolling history (60 s)."""
        timestamp = time.time() - self.session_start_time
        self.signal_history['eda'].append(eda)
        self.signal_history['hr'].append(hr)
        self.signal_history['hrv'].append(hrv)
        self.signal_history['timestamps'].append(timestamp)
        if len(self.signal_history['eda']) > self.max_history_length:
            self.signal_history['eda'].pop(0)
            self.signal_history['hr'].pop(0)
            self.signal_history['hrv'].pop(0)
            self.signal_history['timestamps'].pop(0)

    def record_stress_metric(self, s_instant: float, s_t: float,
                              state: str, dashboard_score: float):
        """Append a computed S_t / state pair to the in-memory history."""
        self.stress_history['s_instant'].append(s_instant)
        self.stress_history['s_t'].append(s_t)
        self.stress_history['state'].append(state)
        self.stress_history['dashboard_score'].append(dashboard_score)

    # ---------- baseline ----------------------------------------------------

    def set_baseline_stats(self, personal_baselines: dict,
                            artifacts_removed: dict):
        """Store baseline statistics after 3-sigma cleaning."""
        self.personal_baselines = personal_baselines
        self.artifacts_removed = artifacts_removed
        print(f"[SESSION] Personal Baselines:")
        print(f"  -> EDA:  {personal_baselines['eda']:.2f} uS")
        print(f"  -> HR:   {personal_baselines['hr']:.2f} BPM")
        print(f"  -> HRV:  {personal_baselines['hrv']:.2f} ms")

    def set_thresholds(self, thresh_mild: float, thresh_high: float):
        """Store classification thresholds locked at baseline end."""
        self.thresh_mild = thresh_mild
        self.thresh_high = thresh_high

    def write_baseline_capture(self, sigma_baseline: float,
                                thresh_mild: float, thresh_high: float,
                                source_label: str = "",
                                hrv_summary: dict = None):
        """
        Persist the frozen baseline by AMENDING the metadata.json that
        the launcher wrote at session start. We don't write a separate
        baseline_*.json file anymore — there's one metadata.json per
        session and the baseline lives inside it.

        Returns the metadata.json path (kept for back-compat with old
        callers that print it).
        """
        baseline_record = {
            "captured_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "baseline_duration_sec": Config.BASELINE_SEC,
            "sample_rate_pipeline_hz": Config.PIPELINE_RATE,
            "personal_baselines": {
                "eda_uS": float(self.personal_baselines.get('eda', 0.0)),
                "hr_bpm": float(self.personal_baselines.get('hr', 0.0)),
                "hrv_ms": float(self.personal_baselines.get('hrv', 0.0)),
            },
            "sigma_baseline": float(sigma_baseline),
            "thresholds": {
                "mild": float(thresh_mild),
                "high": float(thresh_high),
            },
            "hrv_metrics": hrv_summary or {},
            "artifacts_removed": {
                "eda": int(self.artifacts_removed.get('eda', 0)),
                "hr":  int(self.artifacts_removed.get('hr', 0)),
                "hrv": int(self.artifacts_removed.get('hrv', 0)),
            },
            "source": source_label,
            "pipeline_constants": {
                # EMA smoothing removed per PDF — no ema_alpha_* keys here.
                "artifact_sigma_multiplier": Config.ARTIFACT_SIGMA_MULTIPLIER,
                "weight_eda": Config.WEIGHT_EDA,
                "weight_hrv": Config.WEIGHT_HRV,
                "weight_hr":  Config.WEIGHT_HR,
                "thresh_mild_k": Config.THRESH_MILD_K,
                "thresh_high_k": Config.THRESH_HIGH_K,
                "rmssd_window_sec": Config.RMSSD_WINDOW_SEC,
                "hr_window_sec": Config.HR_WINDOW_SEC,
                "hr_compute_interval_sec": Config.HR_COMPUTE_INTERVAL_SEC,
                "rr_max_relative_change": Config.RR_MAX_RELATIVE_CHANGE,
                "hr_sigma_floor_pct": Config.HR_SIGMA_FLOOR_PCT,
                "hrv_sigma_floor_pct": Config.HRV_SIGMA_FLOOR_PCT,
                "eda_phasic_sigma_floor": Config.EDA_PHASIC_SIGMA_FLOOR,
                "eda_phasic_max_us": Config.EDA_PHASIC_MAX_US,
            },
        }
        self.metadata['baseline'] = baseline_record
        # Persist the scenario we learned from Unity's first telemetry
        # packet (if any). Session_review reads this back to label the
        # bottom telemetry panel; harmless when empty (session never had
        # a Unity scene connected).
        if self._observed_scenario:
            self.metadata['scenario'] = self._observed_scenario
        with open(self.metadata_path, 'w', encoding='utf-8') as f:
            json.dump(self.metadata, f, indent=2)
        print(f"[SESSION] Baseline captured into metadata.json")
        return self.metadata_path

    # ---------- shutdown / cleanup helpers ----------------------------------
    #
    # Files now live inside the session folder. The shutdown semantics map to:
    #
    #   SHUTDOWN_SAVE_BOTH    → keep the folder intact, just close handles
    #   SHUTDOWN_DISCARD_LIVE → keep metadata.json (baseline); drop samples.csv,
    #                            diagnostic.csv (live portion), unity_udp.csv
    #   SHUTDOWN_DISCARD_BOTH → delete the WHOLE folder (no files survive)
    #
    # The old per-file delete helpers stay around as no-ops / partial deletes
    # for back-compat with main.py's existing shutdown branches.

    def write_samples_xlsx(self):
        """Write samples.csv next to samples.xlsx, with the first row
        frozen and column widths auto-fitted. Convenience for operators
        who scroll the file in Excel and want the header always visible.

        Called on graceful session end (LIVE -> STOPPED) and on
        Save-Both shutdowns. Skipped silently if openpyxl is missing
        or the CSV does not exist (e.g. discarded session).
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return
        csv_path = self.output_file_path
        if not csv_path or not os.path.exists(csv_path):
            return
        # Flush + close the CSV handle so the read sees every row.
        self._close_samples()
        xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'
        try:
            wb = openpyxl.Workbook(write_only=False)
            ws = wb.active
            ws.title = "samples"
            with open(csv_path, 'r', encoding='utf-8', newline='') as f:
                rows = [line.rstrip('\n').split(',') for line in f if line.strip()]
            if not rows:
                return
            for r in rows:
                # Coerce numeric-looking cells so Excel sorts and plots
                # them correctly; leave the rest as strings.
                ws.append([_coerce_xlsx_cell(c) for c in r])
            # Freeze the first row + autofit column widths.
            ws.freeze_panes = 'A2'
            for col_idx, _ in enumerate(rows[0], start=1):
                letter = get_column_letter(col_idx)
                width = max(len(str(r[col_idx - 1])) for r in rows if col_idx - 1 < len(r))
                ws.column_dimensions[letter].width = min(40, max(8, width + 2))
            # Bold the header row.
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)
            wb.save(xlsx_path)
            print(f"[SESSION] Wrote {os.path.basename(xlsx_path)} "
                  f"(frozen header, autofit columns).")
        except Exception as e:
            print(f"[SESSION] WARN: could not write samples.xlsx: {e}")

    def delete_session_csv(self):
        """Remove samples.csv (and any rotated samples_NNN.csv) from disk."""
        self._close_samples()
        try:
            for fn in os.listdir(self.session_folder):
                if fn.startswith('samples') and fn.endswith('.csv'):
                    p = os.path.join(self.session_folder, fn)
                    os.remove(p)
                    print(f"[SESSION] Deleted: {fn}")
        except OSError as e:
            print(f"[SESSION] WARN: could not delete samples CSVs: {e}")

    def delete_baseline_json(self):
        """
        In the new layout the baseline lives inside metadata.json, not a
        separate file. To 'delete the baseline' we reset its key to None
        and rewrite the file (preserving the intake portion). The
        SHUTDOWN_DISCARD_BOTH path uses delete_session_folder() instead.
        """
        try:
            self.metadata['baseline'] = None
            with open(self.metadata_path, 'w', encoding='utf-8') as f:
                json.dump(self.metadata, f, indent=2)
            print(f"[SESSION] Baseline cleared from metadata.json")
        except OSError as e:
            print(f"[SESSION] WARN: could not clear baseline in metadata.json: {e}")

    def delete_patient_json(self):
        """
        In the old layout this removed a separate patient_*.json. In the
        new layout the intake is part of metadata.json — for the discard
        path it's covered by delete_session_folder(); calling this alone
        is now a no-op so existing main.py shutdown branches don't crash.
        """
        # No-op: the intake lives in metadata.json and is handled by the
        # whole-folder discard path. Kept for back-compat.
        pass

    def delete_live_log(self):
        """No-op: the live-log file was dropped. Kept for back-compat."""
        pass

    def close_live_log(self):
        """No-op: the live-log file was dropped. Kept for back-compat."""
        pass

    def discard_diagnostic_log(self, final: bool = False):
        """
        Close (and optionally truncate-reopen) diagnostic.csv. Replaces
        the per-module discard_log() that acquisition.py and processing.py
        used to expose.

        final=False (default): truncate the file (rewrite the header) so
        the next baseline attempt has a clean slate.
        final=True: close and delete the file with no reopen — used on the
        shutdown-discard path so we don't recreate a file we just removed.
        """
        self._close_diagnostic()
        if final:
            try:
                if os.path.exists(self.diagnostic_path):
                    os.remove(self.diagnostic_path)
                    print(f"[SESSION] Discarded diagnostic.csv (final).")
            except OSError as e:
                print(f"[SESSION] WARN: could not delete diagnostic.csv: {e}")
        else:
            self._open_diagnostic(write_header=True)

    def delete_unity_audit_csv(self):
        """Remove unity_udp.csv (UnityUDPBridge already closed its handle
        via the shutdown path before this is called)."""
        try:
            if os.path.exists(self.unity_udp_path):
                os.remove(self.unity_udp_path)
                print(f"[SESSION] Deleted: unity_udp.csv")
        except OSError as e:
            print(f"[SESSION] WARN: could not delete unity_udp.csv: {e}")

    def delete_session_folder(self):
        """
        Nuke the whole session folder. Used by SHUTDOWN_DISCARD_BOTH —
        the operator explicitly chose to throw away everything.
        Equivalent to rm -rf data/session_<...>/.
        """
        self.close_all()
        try:
            import shutil as _shutil
            if os.path.isdir(self.session_folder):
                _shutil.rmtree(self.session_folder, ignore_errors=False)
                print(f"[SESSION] Deleted session folder: "
                      f"{os.path.basename(self.session_folder)}")
        except OSError as e:
            print(f"[SESSION] WARN: could not delete session folder: {e}")

    # ---------- session summary --------------------------------------------

    def print_session_summary(self):
        """Console summary at LIVE → STOPPED. Compact clinical-style block."""
        states = self.stress_history.get('state') or []
        s_t = self.stress_history.get('s_t') or []
        if not states:
            print("\n[SESSION REVIEW] No live samples recorded — nothing to summarize.\n")
            return
        n = len(states)
        rate = float(Config.PIPELINE_RATE)
        live_sec = n / rate
        n_calm = sum(1 for s in states if s == 'calm')
        n_stress = sum(1 for s in states if s == 'stressed')
        n_ultra = sum(1 for s in states if s == 'ultra_stressed')

        def _pct(k):
            return 100.0 * k / n if n else 0.0

        mean_st = sum(s_t) / len(s_t) if s_t else 0.0
        max_st = max(s_t) if s_t else 0.0

        print()
        print("=" * 56)
        print("  SESSION REVIEW")
        print("=" * 56)
        print(f"  Patient        : {self.patient_info}")
        print(f"  Session folder : {os.path.basename(self.session_folder)}")
        print(f"  Live duration  : {live_sec:6.1f} s  ({n} samples)")
        print(f"  Time in CALM   : {n_calm/rate:6.1f} s  ({_pct(n_calm):5.1f}%)")
        print(f"  Time in STRESS : {n_stress/rate:6.1f} s  ({_pct(n_stress):5.1f}%)")
        print(f"  Time in ULTRA  : {n_ultra/rate:6.1f} s  ({_pct(n_ultra):5.1f}%)")
        print(f"  Mean S_t       : {mean_st:+6.2f}")
        print(f"  Max  S_t       : {max_st:+6.2f}")
        print(f"  Samples CSV    : {os.path.basename(self.samples_path)}")
        print(f"  Metadata JSON  : metadata.json")
        print("=" * 56)
        print()

    # ---------- back-compat / dashboard helpers ----------------------------

    def update_phase_baseline(self, is_baseline_complete: bool):
        """Called when processing indicates baseline is complete."""
        if is_baseline_complete and self.phase == "BASELINE":
            self.phase = "LIVE"
            self.baseline_end_time = time.time()
            elapsed = self.baseline_end_time - self.session_start_time
            print(f"\n[SESSION] BASELINE COMPLETE at {elapsed:.1f}s")
            print(f"[SESSION] Phase transition: BASELINE -> LIVE\n")

    def get_session_duration_sec(self) -> float:
        return time.time() - self.session_start_time

    def get_session_duration_str(self) -> str:
        elapsed = self.get_session_duration_sec()
        m = int(elapsed // 60)
        s = int(elapsed % 60)
        return f"{m:02d}:{s:02d}"

    def get_current_state_summary(self) -> dict:
        """Return current session state as a dict (used by some dashboard code)."""
        return {
            'patient_name': self.patient_name,
            'patient_id': self.patient_id,
            'patient_info': self.patient_info,
            'phase': self.phase,
            'duration': self.get_session_duration_str(),
            'duration_sec': self.get_session_duration_sec(),
            'baseline_complete': (self.phase == "LIVE"),
            'personal_baselines': self.personal_baselines,
            'thresh_mild': self.thresh_mild,
            'thresh_high': self.thresh_high,
            'artifacts_removed': self.artifacts_removed,
            'signal_count': len(self.signal_history['eda']),
            'stress_events_count': len(self.stress_history['s_t']),
            'current_state': self.stress_history['state'][-1] if self.stress_history['state'] else None,
            'current_s_t': self.stress_history['s_t'][-1] if self.stress_history['s_t'] else None,
            'current_dashboard_score': self.stress_history['dashboard_score'][-1] if self.stress_history['dashboard_score'] else None,
            'output_file': self.output_file_path,
            'session_folder': self.session_folder,
        }

    # log_live_line was removed (the standalone live_log.txt was dropped).
    # main.py used to call this; it now just echoes to stdout directly.
    def log_live_line(self, line: str):
        """Back-compat: the standalone live_log.txt file was dropped.
        We just echo to stdout so the operator still sees the per-second
        transcript in the terminal. To regenerate a paste-friendly text
        file post-hoc, run `python src/session_review.py <session-folder>`
        which can derive it from samples.csv."""
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"{ts} | {line}")


if __name__ == "__main__":
    print("\n=== SESSION MANAGER TEST ===\n")
    session = SessionManager()
    for i in range(10):
        session.record_raw_sample(5.0 + i * 0.1, 75.0 + i * 0.5, 40.0)
        session.record_stress_metric(0.5, 0.3, "calm", 25.0)
    session.set_baseline_stats(
        {'eda': 5.0, 'hr': 75.0, 'hrv': 40.0},
        {'eda': 2, 'hr': 0, 'hrv': 1},
    )
    session.set_thresholds(1.28, 2.33)
    summary = session.get_current_state_summary()
    print("\nCurrent State Summary:")
    for key, value in summary.items():
        print(f"  {key}: {value}")
    session.close_all()
